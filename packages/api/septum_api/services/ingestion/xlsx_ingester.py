from __future__ import annotations

"""XLSX spreadsheet ingester using openpyxl.

This ingester is responsible for:
    - Reading the encrypted XLSX bytes from disk.
    - Decrypting them in memory using the shared AES-256-GCM utilities.
    - Extracting plain text by iterating over sheets and rows.
    - Returning an :class:`IngestionResult` with the concatenated text and
      lightweight, non-PII metadata (e.g., sheet names, row counts).
"""

import asyncio
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook  # type: ignore[import]

from ...utils.crypto import decrypt
from .base import BaseIngester, IngestionResult


class XlsxIngester(BaseIngester):
    """Ingests encrypted XLSX spreadsheets and extracts their textual content."""

    async def ingest(
        self,
        file_path: Path,
        *,
        mime_type: str,
        file_format: str,
    ) -> IngestionResult:
        """Ingest the XLSX at ``file_path`` and return extracted content."""

        return await asyncio.to_thread(
            self._ingest_sync,
            file_path,
            mime_type,
            file_format,
        )

    def _ingest_sync(
        self,
        file_path: Path,
        mime_type: str,
        file_format: str,
    ) -> IngestionResult:
        """Synchronous part of XLSX ingestion, run in a worker thread."""

        encrypted_bytes = file_path.read_bytes()
        xlsx_bytes = decrypt(encrypted_bytes)

        workbook = load_workbook(BytesIO(xlsx_bytes), data_only=True, read_only=True)

        sheet_texts: List[str] = []
        sheet_metadata: List[Dict[str, Any]] = []

        for sheet in workbook.worksheets:
            rows_text: List[str] = []
            row_count = 0
            header_cells: List[str] | None = None

            for row in sheet.iter_rows(values_only=True):
                row_count += 1
                raw_values = [value for value in row if value is not None]
                if not raw_values:
                    continue

                if header_cells is None:
                    # First non-empty row is treated as the header; replace the
                    # original header text with generic, non-PII column labels.
                    header_cells = [f"COLUMN_{idx + 1}" for idx, _ in enumerate(raw_values)]
                    rows_text.append("\t".join(header_cells))
                    continue

                cells = [str(value) for value in raw_values]
                if cells:
                    rows_text.append("\t".join(cells))

            sheet_text = "\n".join(rows_text)
            if sheet_text:
                header_suffix = ""
                if header_cells:
                    header_suffix = " — columns: " + " | ".join(header_cells)
                sheet_texts.append(f"# Sheet: {sheet.title}{header_suffix}\n{sheet_text}")

            sheet_metadata.append(
                {
                    "sheet_name": sheet.title,
                    "row_count": row_count,
                    "column_count": len(header_cells) if header_cells is not None else 0,
                }
            )

        full_text = "\n\n".join(sheet_texts)

        metadata: Dict[str, Any] = {
            "sheet_count": len(workbook.worksheets),
            "sheets": sheet_metadata,
            "mime_type": mime_type,
            "file_format": file_format,
        }

        return IngestionResult(text=full_text, metadata=metadata)

